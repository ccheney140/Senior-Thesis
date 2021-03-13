#Libraries used
from openpyxl import load_workbook
import os
from pathlib import Path
import pandas as pd

wb = load_workbook("Data.xlsx")  # Work Book Name
ws = wb.get_sheet_by_name('Sheet1')  # Work Sheet Name
column = ws['C']  # Column Name to search
column_list = [column[x].value for x in range(len(column))]

#Converts the excell cell data into lists
kAt = column_list[0].split(',') 
kBt = column_list[1].split(',') 
kCt = column_list[2].split(',')

#Creates lists to use
kA=[]
kB=[]
kC=[]

#Gets correct data from the k_T lists and adds it to the k_ lists above
for i in range(len(kAt)):
    if(kAt[i][0]=='A' and kAt[i][1]=='G'): #Checks if the value starts with "AG", if it does it adds it to the k_ list
        kA.append(kAt[i])

for i in range(len(kBt)):
    if(kBt[i][0]=='A' and kBt[i][1]=='G'):
        kB.append(kBt[i])

for i in range(len(kCt)):
    if(kCt[i][0]=='A' and kCt[i][1]=='G'):
        kC.append(kCt[i])

#Finds all of the files in the folder
x = [i[2] for i in os.walk('.')]
y=[]
for t in x:
    for f in t:
        y.append(f)
z=[]
for i in y:
    if i.find('.genes.fna')!=-1: #Gets all .genes.fna files
        z.append(i.replace('.genes.fna','')) 

#Converts the files into easily read dictionaries
def getFileData(filename):
    txt = Path("C:/Users/chene/Desktop/cf/Data/"+filename+'.genes.fna').read_text() #opens the file
    #txt=txt.replace('\n','')
    txt=txt.split('>') #Splits the file by the ">" character
    data=txt
    tempdata=[]

    for i in data:
        if (i!=' ' and i !=''): #Removes unnecessary values
            tempdata.append(i)
    data=tempdata
    data1={} #Creates a dictionary
    for i in range(len(data)): 
        temp=data[i]
        temp=temp[0:temp.find(' ')] #Cuts the file using the " " character

        temp2=data[i]
        temp2=temp2.replace(temp,'') #Removes a substring from the string

        data1[temp]=temp2

    data2={} #Creates a dictionary
    fn=filename[0:filename.find('_')] #Modifies the file name
    data2[fn]=data1 #Saves all of the file data to the dictionary
    return data2 #Returns the final data

allData={}

def addData(filename): #Adds the file data to a larger dictionary
    od=getFileData(filename)
    addkey=list(od.keys())[0]
    allData[addkey] = od[addkey]

for i in z: #Goes through every file and adds them to the allData dictionary
    addData(str(i))
 
#Goes through all of the excel data and finds the matching information from the file data. Saving that information to a list
kAd1=[]

for i in kA:
    first = i[0:i.find('_')]
    second = i[i.find('_')+1:]
    kAd1.append('>'+second+allData[first][second])

kBd1=[]

for i in kB:
    first = i[0:i.find('_')]
    second = i[i.find('_')+1:]
    kBd1.append('>'+second+allData[first][second])

kCd1=[]

for i in kC:
    first = i[0:i.find('_')]
    second = i[i.find('_')+1:]
    kCd1.append('>'+second+allData[first][second])

#Creates a text file using the gathered data
with open('A.txt', 'w') as f:
    for item in kAd1:
        f.write(item)

with open('B.txt', 'w') as f:
    for item in kBd1:
        f.write(item)

with open('C.txt', 'w') as f:
    for item in kCd1:
        f.write(item)