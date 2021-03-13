import pandas as pd
from openpyxl import load_workbook
to_read=[]

def readBigBoy():
    global to_read
    to_read=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO']
    
    wb = load_workbook('genome_assembly_summary_20180718_SNC.xlsx')  # Work Book
    ws = wb.get_sheet_by_name('genome_assembly_summary_2018071')  # Work Sheet
    
    whyiseverythinghere=[]
    for i in to_read:
        column = ws[i]  # Column
        c2 = [str(column[x].value) for x in range(len(column))]
        whyiseverythinghere.append(c2)
    wb.close()

    return whyiseverythinghere


def doStuff(start,end,out):
    global to_read
    df = pd.read_excel(open(start, 'rb'),sheet_name='Sheet1')
    c1 = df['Tree node ID'].tolist()
    for i in range(len(c1)):
        c1[i]=c1[i].replace(' ','_')

    wb = load_workbook(end)  # Work Book
    ws = wb.get_sheet_by_name('My sheet')  # Work Sheet
    column = ws['A']  # Column
    c2 = [str(column[x].value) for x in range(len(column))]
    column = ws['B']  # Column
    c3 = [str(column[x].value) for x in range(len(column))]
    wb.close()
    c_dict={}
    c3_3=[]

    for i in range(len(c2)):
        val=c2[i]
        val=val.replace('(-)','')
        val=val.replace('(+)','')
        val1=val.find('(Ga')
        val2=val.find(')')
        val=val[val1+1:val2]
        c3_3.append(val)
        c_dict[val]=c3[i]

    c4=[]
    for i in c1:
        c4.append(c_dict[i])
    
    c5=[]
    for i in c4:
        val=i.replace('(Screened)','')
        val1=val.find('AG')
        val2=val.find(']')
        val=val[val1:val2]
        c5.append(val.replace(' ',''))

    realbigboy = readBigBoy()
   
    everything=[]

    inc=0
    for i in c5:
        val = realbigboy[0].index(i)
        temp=[]
        temp.append(c1[inc])
        inc+=1
        for j in realbigboy:
            temp.append(j[val])
        everything.append(temp)
    
    df = pd.DataFrame()
    inc=0
    temp=[]
    for j in range(len(to_read)):
        temp=[]
        for i in everything:
            temp.append(i[j])
        df['PlaceHolder'+str(j)] = temp

    df.to_excel(out, index = False)

doStuff('KaiC Tree Node Data.xlsx','C.xlsx','C2_output.xlsx')
doStuff('KaiB Tree Node Data.xlsx','B.xlsx','B2_output.xlsx')