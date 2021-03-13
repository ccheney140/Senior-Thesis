import xlsxwriter 
from openpyxl import load_workbook
import os
from pathlib import Path
import pandas as pd

wb = load_workbook("Data.xlsx")  # Work Book Name
ws = wb.get_sheet_by_name('Sheet1')  # Work Sheet Name
column = ws['C']  # Column Name to search
column_list = [column[x].value for x in range(len(column))]

#Converts the excell cell data into lists
kAt = column_list[0].split(',') 
kBt = column_list[1].split(',') 
kCt = column_list[2].split(',')

#Creates lists to use
kA=[]
kB=[]
kC=[]

#Gets correct data from the k_T lists and adds it to the k_ lists above
for i in range(len(kAt)):
    if(kAt[i][0]=='A' and kAt[i][1]=='G'): #Checks if the value starts with "AG", if it does it adds it to the k_ list
        kA.append(kAt[i])

for i in range(len(kBt)):
    if(kBt[i][0]=='A' and kBt[i][1]=='G'):
        kB.append(kBt[i])

for i in range(len(kCt)):
    if(kCt[i][0]=='A' and kCt[i][1]=='G'):
        kC.append(kCt[i])

#Finds all of the files in the folder
x = [i[2] for i in os.walk('.')]
y=[]
for t in x:
    for f in t:
        y.append(f)
z=[]
for i in y:
    if i.find('.genes.fna')!=-1: #Gets all .genes.fna files
        z.append(i.replace('.genes.fna','')) 

#Converts the files into easily read dictionaries
def getFileData(filename):
    txt = Path("C:/Users/chene/Desktop/cf/Data/"+filename+'.genes.fna').read_text() #opens the file
    #txt=txt.replace('\n','')
    txt=txt.split('>') #Splits the file by the ">" character
    data=txt
    tempdata=[]

    for i in data:
        if (i!=' ' and i !=''): #Removes unnecessary values
            tempdata.append(i)
    data=tempdata
    data1={} #Creates a dictionary
    for i in range(len(data)): 
        temp=data[i]
        temp=temp[0:temp.find(' ')] #Cuts the file using the " " character

        temp2=data[i]
        temp2=temp2.replace(temp,'') #Removes a substring from the string
        temp2= temp2[0:temp2.find(']')+1]

        data1[temp]=temp2

    data2={} #Creates a dictionary
    fn=filename[0:filename.find('_')] #Modifies the file name
    data2[fn]=data1 #Saves all of the file data to the dictionary
    return data2 #Returns the final data

allData={}

def addData(filename): #Adds the file data to a larger dictionary
    od=getFileData(filename)
    addkey=list(od.keys())[0]
    allData[addkey] = od[addkey]

for i in z: #Goes through every file and adds them to the allData dictionary
    addData(str(i))
 
#Goes through all of the excel data and finds the matching information from the file data. Saving that information to a list
kAd1=[]

for i in kA:
    first = i[0:i.find('_')]
    second = i[i.find('_')+1:]
    s1=('>'+second+allData[first][second])
    kAd1.append([s1[0:s1.find('[')],s1[s1.find('['):]])
    #kAd2.append(s1[s1.find('['):])

kCd1=[]

for i in kC:
    first = i[0:i.find('_')]
    second = i[i.find('_')+1:]
    s1=('>'+second+allData[first][second])
    kCd1.append([s1[0:s1.find('[')],s1[s1.find('['):]])
    #kAd2.append(s1[s1.find('['):])

kBd1=[]

for i in kB:
    first = i[0:i.find('_')]
    second = i[i.find('_')+1:]
    s1=('>'+second+allData[first][second])
    kBd1.append([s1[0:s1.find('[')],s1[s1.find('['):]])
    #kAd2.append(s1[s1.find('['):])

workbook = xlsxwriter.Workbook('A.xlsx') 
  
# By default worksheet names in the spreadsheet will be  
# Sheet1, Sheet2 etc., but we can also specify a name. 
worksheet = workbook.add_worksheet("My sheet") 
  
# Some data we want to write to the worksheet. 
scores = (kAd1) 
  
# Start from the first cell. Rows and 
# columns are zero indexed. 
row = 0
col = 0
  
# Iterate over the data and write it out row by row. 
for name, score in (scores): 
    worksheet.write(row, col, name) 
    worksheet.write(row, col + 1, score) 
    row += 1
  
workbook.close() 

workbook = xlsxwriter.Workbook('B.xlsx') 
  
# By default worksheet names in the spreadsheet will be  
# Sheet1, Sheet2 etc., but we can also specify a name. 
worksheet = workbook.add_worksheet("My sheet") 
  
# Some data we want to write to the worksheet. 
scores = (kBd1) 
  
# Start from the first cell. Rows and 
# columns are zero indexed. 
row = 0
col = 0
  
# Iterate over the data and write it out row by row. 
for name, score in (scores): 
    worksheet.write(row, col, name) 
    worksheet.write(row, col + 1, score) 
    row += 1
  
workbook.close() 

workbook = xlsxwriter.Workbook('C.xlsx') 
  
# By default worksheet names in the spreadsheet will be  
# Sheet1, Sheet2 etc., but we can also specify a name. 
worksheet = workbook.add_worksheet("My sheet") 
  
# Some data we want to write to the worksheet. 
scores = (kCd1) 
  
# Start from the first cell. Rows and 
# columns are zero indexed. 
row = 0
col = 0
  
# Iterate over the data and write it out row by row. 
for name, score in (scores): 
    worksheet.write(row, col, name) 
    worksheet.write(row, col + 1, score) 
    row += 1
  
workbook.close() 
